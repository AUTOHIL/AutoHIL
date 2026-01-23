import re
from dataclasses import dataclass, field
from typing import Dict, Optional, Tuple, Set, List
import os
import json
import csv 
from elftools.elf.elffile import ELFFile
import openpyxl
from tqdm import tqdm


# =========================
# 1) Output writers (Excel / CSV)
# =========================

class ExcelWriter:
    HEADER = [
        "TOP LEVEL VARIABLE", "TOP LEVEL SIZE", "MEMBER PATH",
        "ELEMENT SIZE", "ELEMENT DATA TYPE", "FILE NAME", "LINE"
    ]

    def __init__(self) -> None:
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "DWARF Parser Results"
        self.row = 1
        for col, title in enumerate(self.HEADER, start=1):
            self.ws.cell(row=self.row, column=col, value=title)
        self.row += 1
        self._seen: Set[Tuple] = set()  # deduplication key

    def write(self,
              top_level_var: str,
              top_level_size: Optional[int],
              member_path: str,
              element_size: Optional[int],
              element_type: Optional[str],
              file_name: str,
              line: str
             ) -> None:
        
        dedup_key = (top_level_var, member_path, element_size, element_type, file_name, line)
        if dedup_key in self._seen:
            return
        self._seen.add(dedup_key)

        self.ws.cell(row=self.row, column=1, value=top_level_var)
        self.ws.cell(row=self.row, column=2, value=top_level_size if top_level_size is not None else "")
        self.ws.cell(row=self.row, column=3, value=member_path)
        self.ws.cell(row=self.row, column=4, value=element_size if element_size is not None else "")
        self.ws.cell(row=self.row, column=5, value=element_type or "")
        self.ws.cell(row=self.row, column=6, value=file_name)
        self.ws.cell(row=self.row, column=7, value=line)
        self.row += 1

    def save(self, path: str) -> None:
        self.wb.save(path)

class CsvWriter:
    HEADER = [
        "TOP LEVEL VARIABLE", "TOP LEVEL SIZE", "MEMBER PATH",
        "ELEMENT SIZE", "ELEMENT DATA TYPE", "FILE NAME", "LINE"
    ]

    def __init__(self) -> None:
        self.rows: List[List[str]] = []
        self.rows.append(self.HEADER)
        self._seen: Set[Tuple] = set()  # deduplication key

    def write(self,
              top_level_var: str,
              top_level_size: Optional[int],
              member_path: str,
              element_size: Optional[int],
              element_type: Optional[str],
              file_name: str,
              line: str
             ) -> None:
        
        dedup_key = (top_level_var, member_path, element_size, element_type, file_name, line)
        if dedup_key in self._seen:
            return
        self._seen.add(dedup_key)
        
        row_data = [
            top_level_var,
            str(top_level_size) if top_level_size is not None else "",
            member_path,
            str(element_size) if element_size is not None else "",
            element_type or "",
            file_name,
            line
        ]
        self.rows.append(row_data)

    def save(self, path: str) -> None:
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerows(self.rows)
        except PermissionError:
            print(f"Error: Failed to write {path}; the file may be open in another program.")
        except Exception as e:
            print(f"Unknown error occurred while writing CSV: {e}")

# =========================
# 2) DWARF utilities
# =========================

def _get_die_name(die) -> Optional[str]:
    if not die or "DW_AT_name" not in die.attributes:
        return None
    val = die.attributes["DW_AT_name"].value
    return val.decode("utf-8", errors="ignore") if isinstance(val, bytes) else str(val)

def _get_die_byte_size(die) -> int:
    if not die or "DW_AT_byte_size" not in die.attributes:
        return 0
    val = die.attributes["DW_AT_byte_size"].value
    return int(val) if isinstance(val, int) else 0

def _data_type_desc(die) -> str:
    if not die:
        return "unknown"
    tag = die.tag
    name = _get_die_name(die)
    if tag == "DW_TAG_base_type":
        if not name:
            return "unknown base type"
        mapping = {
            "int": "signed int",
            "unsigned int": "unsigned int",
            "char": "signed char",
            "unsigned char": "unsigned char",
            "short": "signed short int",
            "unsigned short": "unsigned short int",
            "long": "signed long int",
            "unsigned long": "unsigned long int",
            "long long": "signed long long int",
            "unsigned long long": "unsigned long long int",
            "float": "float",
            "double": "double",
        }
        return mapping.get(name, name)
    if tag == "DW_TAG_enumeration_type":
        return "enum"
    if tag == "DW_TAG_structure_type":
        return "struct"
    if tag == "DW_TAG_union_type":
        return "union"
    if tag == "DW_TAG_array_type":
        return "array"
    if tag == "DW_TAG_pointer_type":
        return "pointer"
    if tag == "DW_TAG_typedef":
        if "DW_AT_type" in die.attributes:
            try:
                return _data_type_desc(die.get_DIE_from_attribute("DW_AT_type"))
            except Exception:
                pass
        return name or "typedef"
    return name or "unknown"


def _follow_type(die) -> Tuple[Optional[object], Optional[str]]:
    visited = set()
    cur = die
    last_name = None
    while cur and cur.offset not in visited:
        visited.add(cur.offset)
        nm = _get_die_name(cur)
        if nm:
            last_name = nm
        if cur.tag in (
            "DW_TAG_structure_type", "DW_TAG_union_type",
            "DW_TAG_array_type", "DW_TAG_base_type",
            "DW_TAG_enumeration_type"
        ):
            return cur, last_name
        if cur.tag == "DW_TAG_typedef" and "DW_AT_type" in cur.attributes:
            try:
                cur = cur.get_DIE_from_attribute("DW_AT_type")
                continue
            except Exception:
                break
        if "DW_AT_type" in cur.attributes:
            try:
                cur = cur.get_DIE_from_attribute("DW_AT_type")
            except Exception:
                break
        else:
            break
    return cur, last_name


# =========================
# 3) DWARF traversal + write Excel
# =========================

class DwarfEmitter:
    """Flatten types from ELF variables filtered by file_whitelist and write to Excel/CSV"""
    
    def __init__(self, elf_path: str, file_whitelist: Set[str], xw: (ExcelWriter | CsvWriter)) -> None:
        self.elf_path = elf_path
        self.file_whitelist = file_whitelist
        self.xw = xw 

    def _is_valid_variable_name(self, name: str) -> bool:
        """
        Helper to filter out compiler/linker placeholder symbols.
        """
        if not name:
            return False
        # filter out C++/internal symbols (e.g., _Z..., .L..., __static_...)
        if name.startswith('_') or name.startswith('.'):
            return False
        # filter other odd symbols
        if '$' in name or '@' in name or 'vftable' in name:
            return False
        return True

    def run_with_cu_file_filter(self) -> None:
        with open(self.elf_path, "rb") as f:
            elf = ELFFile(f)
            if not elf.has_dwarf_info():
                print("ELF file is missing DWARF debug information")
                return
            dwarfinfo = elf.get_dwarf_info()

        processed_vars = 0

        # iterate over all compilation units (CUs)
        for cu in tqdm(dwarfinfo.iter_CUs(), desc="Parsing compilation units (CU)"):
            
            # 1. get CU's main file name (fallback)
            cu_file_name = "unknown_cu_file"
            try:
                top_die = cu.get_top_DIE()
                if "DW_AT_name" in top_die.attributes:
                    cu_file_name = top_die.attributes["DW_AT_name"].value.decode(errors="ignore")
            except Exception:
                # If the CU doesn't even have a top_die, skip it
                continue 
            
            if os.name == "posix":
                cu_file_name = cu_file_name.replace("\\", "/")
            cu_basename = os.path.basename(cu_file_name)
            
            # 2. <--- Filter #1: file whitelist (case-insensitive) ---
            #    Skip the CU entirely if its basename is not in the whitelist
            if cu_basename.lower() not in self.file_whitelist:
                continue

            # --- Whitelist passed; begin iterating DIEs in this CU ---

            # iterate DIEs in this CU
            for die in cu.iter_DIEs():
                # only consider top-level variables
                if not (die.tag == "DW_TAG_variable" and "DW_AT_name" in die.attributes):
                    continue

                var_name = _get_die_name(die)
                
                # 3. <--- Filter #2: variable name (filter placeholders) ---
                if not self._is_valid_variable_name(var_name):
                    continue
                
                # --- All filters passed; begin processing ---
                processed_vars += 1

                if 'DW_AT_type' not in die.attributes:
                    continue

                try:
                    type_die = die.get_DIE_from_attribute('DW_AT_type')
                    actual_type, _ = _follow_type(type_die)
                    if not actual_type:
                        continue
                    
                    top_level_size = _get_die_byte_size(actual_type)

                    self._traverse(
                        die=actual_type,
                        top_level_var=var_name,
                        top_level_size=top_level_size,
                        current_path=var_name,
                        file_name=cu_basename, # <--- write reliable CU file name
                        visited=set()
                    )
                except Exception as e:
                    print(f"Error while processing variable {var_name} (in {cu_basename}): {e}")
                    continue
            
        print(f"Successfully processed {processed_vars} top-level variables in the whitelist")

    def run_with_file_table_filter(self) -> None:
        with open(self.elf_path, "rb") as f:
            elf = ELFFile(f)
            if not elf.has_dwarf_info():
                print("ELF file is missing DWARF debug information")
                return
            dwarfinfo = elf.get_dwarf_info()

        print(f"(Current runtime OS: {'Windows' if os.name == 'nt' else 'Linux/POSIX'})")
        processed_vars = 0

        # iterate over all compilation units (CUs)
        for cu in tqdm(dwarfinfo.iter_CUs(), desc="Parsing compilation units (CU)"):
            
            # 1. get CU's main file name (used as fallback)
            cu_file_name = "unknown_cu_file"
            cu_basename = "unknown_cu_file"
            try:
                top_die = cu.get_top_DIE()
                if "DW_AT_name" in top_die.attributes:
                    cu_file_name = top_die.attributes["DW_AT_name"].value.decode(errors="ignore")
                
                if os.name == "posix":
                    cu_file_name = cu_file_name.replace("\\", "/")
                cu_basename = os.path.basename(cu_file_name)
                
            except Exception:
                continue # skip invalid CU
            
            # 2. obtain the file table for this CU (copied from reference code)
            line_prog = dwarfinfo.line_program_for_CU(cu)
            file_table: Dict[int, str] = {}
            if line_prog:
                for i, entry in enumerate(line_prog.header['file_entry']):
                    file_table[i+1] = entry.name.decode(errors='ignore') # DWARF index is 1-based

            # iterate DIEs in this CU
            for die in cu.iter_DIEs():
                # only consider top-level variables
                if not (die.tag == "DW_TAG_variable" and "DW_AT_name" in die.attributes):
                    continue

                var_name = _get_die_name(die)
                
                # 3. <--- Filter #2: variable name (filter placeholders) ---
                if not self._is_valid_variable_name(var_name):
                    continue
                
                
                # 4. extract declaration line number
                line = '?'
                if 'DW_AT_decl_line' in die.attributes:
                    line = str(die.attributes['DW_AT_decl_line'].value)
                
                # 5. extract declaration file
                file_basename = cu_basename # default to the CU's main file
                if 'DW_AT_decl_file' in die.attributes:
                    file_idx = die.attributes['DW_AT_decl_file'].value
                    file_name_from_table = file_table.get(file_idx, cu_basename)
                    if os.name == "posix":
                        file_name_from_table = file_name_from_table.replace("\\", "/")
                    file_basename = os.path.basename(file_name_from_table) # prefer .h, fallback to .c
                
                # 1. <--- Filter #1: file whitelist (case-insensitive) ---
                if file_basename.lower() not in self.file_whitelist:
                    continue
                
                # --- All filters passed; begin processing ---
                processed_vars += 1

                if 'DW_AT_type' not in die.attributes:
                    continue

                try:
                    type_die = die.get_DIE_from_attribute('DW_AT_type')
                    actual_type, _ = _follow_type(type_die)
                    if not actual_type:
                        continue
                    
                    top_level_size = _get_die_byte_size(actual_type)

                    # 6. pass line number to _traverse
                    self._traverse(
                        die=actual_type,
                        top_level_var=var_name,
                        top_level_size=top_level_size,
                        current_path=var_name,
                        file_name=file_basename, # <--- write reliable CU file name
                        line=line,               # <--- pass line number
                        visited=set()
                    )
                except Exception as e:
                    print(f"Error while processing variable {var_name} (in {cu_basename}): {e}")
                    continue
            
        print(f"Successfully processed {processed_vars} top-level variables in the whitelist")

    def _traverse(self,
                  die,
                  top_level_var: str,        
                  top_level_size: int,       
                  current_path: str,         
                  file_name: str,
                  line: str,
                  visited: Set[int]) -> None:
        
        if not die or getattr(die, "offset", None) in visited:
            return
        visited.add(die.offset)
        tag = die.tag
        if tag == "DW_TAG_array_type":
            elem_die = None
            if "DW_AT_type" in die.attributes:
                try:
                    elem_die, _ = _follow_type(die.get_DIE_from_attribute("DW_AT_type"))
                except Exception:
                    elem_die = None
            new_path = f"{current_path}[i]"
            if elem_die:
                self._traverse(elem_die, top_level_var, top_level_size, new_path, file_name, line, visited=set(visited))
            else:
                self.xw.write(top_level_var, top_level_size, new_path, 0, "array element", file_name, line)
            return
        if tag in ("DW_TAG_structure_type", "DW_TAG_union_type"):
            for ch in die.iter_children():
                if ch.tag != "DW_TAG_member":
                    continue
                mname_attr = ch.attributes.get("DW_AT_name")
                mname = (mname_attr.value.decode(errors="ignore") if mname_attr and isinstance(mname_attr.value, bytes)
                         else (str(mname_attr.value) if mname_attr else "anon_member"))
                full = f"{current_path}.{mname}"
                if "DW_AT_type" in ch.attributes:
                    try:
                        mtype = ch.get_DIE_from_attribute("DW_AT_type")
                        actual, _ = _follow_type(mtype)
                        self._traverse(actual, top_level_var, top_level_size, full, file_name, line, visited=set(visited))
                    except Exception:
                        msize = _get_die_byte_size(ch)
                        mdesc = _data_type_desc(ch)
                        self.xw.write(top_level_var, top_level_size, full, msize, mdesc, file_name, line)
                else:
                    self.xw.write(top_level_var, top_level_size, full, 0, "unknown", file_name, line)
            return
        if tag == "DW_TAG_typedef":
            if "DW_AT_type" in die.attributes:
                try:
                    target, _ = _follow_type(die.get_DIE_from_attribute("DW_AT_type"))
                    self._traverse(target, top_level_var, top_level_size, current_path, file_name, line, visited=set(visited))
                except Exception:
                    pass
            return
        if tag in ("DW_TAG_enumeration_type", "DW_TAG_base_type"):
            size = _get_die_byte_size(die)
            desc = _data_type_desc(die)
            self.xw.write(top_level_var, top_level_size, current_path, size, desc, file_name, line)
            return


# =========================
# 4) Main assembly flow
# =========================

def main() -> None:
    
    # ===================================================================
    # Configuration
    # ===================================================================
    USE_CSV_WRITER = True
    elf_file = "Application.elf"
    json_file = "./reqAugmentation/output/CanMgr/cTarFiles.json"
    # ===================================================================
    
    # 1. load whitelist from JSON
    try:
        with open(json_file, "r") as f:
            file_list_from_json = json.loads(f.read())
        
        # convert to basenames and lowercase for case-insensitive matching
        FILE_WHITELIST: Set[str] = set(map(lambda x: os.path.basename(x).lower(), file_list_from_json))
        FILE_WHITELIST = {basename for basename in FILE_WHITELIST 
                                                    if basename.startswith("can") or basename.startswith("com")
                                                }  # remove empty strings
        
    except FileNotFoundError:
        print(f"Error: JSON config file not found: {json_file}")
        print("An empty file whitelist will be used.")
        FILE_WHITELIST = set()
    except json.JSONDecodeError:
        print(f"Error: Failed to parse JSON file: {json_file}")
        print("An empty file whitelist will be used.")
        FILE_WHITELIST = set()


    print("Starting full DWARF parsing...")
    print(f"Target ELF: {elf_file}")

    # 2. choose writer based on configuration
    if USE_CSV_WRITER:
        print(f"Output format: CSV (case-insensitive whitelist filtering enabled)")
        xw = CsvWriter()
        output_filename = "dwarf_filter_analysis.csv"
    else:
        print(f"Output format: Excel (xlsx) (case-insensitive whitelist filtering enabled)")
        xw = ExcelWriter()
        output_filename = "dwarf_filter_analysis.xlsx"


    # 3. run the DWARF emitter
    DwarfEmitter(elf_file, FILE_WHITELIST, xw).run_with_file_table_filter()

    # 4. Save results
    xw.save(output_filename)
    print(f"Parsing complete. Results have been saved to {output_filename}")


if __name__ == "__main__":
    main()
